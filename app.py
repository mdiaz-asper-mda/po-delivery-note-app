import os
import io
import re
import json
import tempfile
from copy import deepcopy
from datetime import datetime
from io import StringIO

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from pypdf import PdfReader
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from copy import copy


TEMPLATE_PATH = "TEMPLATE_PO_APP_260504.xlsx"

st.set_page_config(page_title="Invoice and Delivery Note Tool", layout="wide")
st.title("Invoice and Delivery Note Tool")

if "GEMINI_API_KEY" in st.secrets:
    api_key = st.secrets["GEMINI_API_KEY"]
else:
    load_dotenv()
    api_key = os.getenv("GEMINI_API_KEY")

if not api_key:
    st.error("GEMINI_API_KEY not found. Check Streamlit secrets.")
    st.stop()

client = genai.Client(api_key=api_key)


PRODUCT_CATALOG = [
    {
        "key": "excitatory_cells",
        "aliases": ["excitatory", "quick neuron excitatory", "quick-neuron excitatory", "ex-sev-hc-cw50065"],
        "title": "Quick Neuron Excitatory Human iPSC derived Neurons",
        "reagents": [
            "Cryopreserved Cells (EX-SeV-HC-CW50065)",
            "Comp N (2x840μl)",
            "Comp G2 (60μl)",
            "Comp P (50μl)",
        ],
    },
    {
        "key": "cholinergic_cells",
        "aliases": ["cholinergic", "quick neuron cholinergic", "quick-neuron cholinergic", "ch-sev-hc-cw50065"],
        "title": "Quick Neuron Cholinergic Human iPSC derived Neurons",
        "reagents": [
            "Cryopreserved Cells (CH-SeV-HC-CW50065)",
            "Comp N1 (830μl)",
            "Comp A (80μl)",
            "Comp P (50μl)",
        ],
    },
    {
        "key": "dopaminergic_cells",
        "aliases": ["dopaminergic", "quick neuron dopaminergic", "quick-neuron dopaminergic", "da-mrna-hc-cw50065"],
        "title": "Quick Neuron Dopaminergic Human iPSC derived Neurons",
        "reagents": [
            "Cryopreserved Cells (DA-mRNA-HC-CW50065)",
            "Comp N (840μl)",
            "Comp P (50μl)",
            "Comp D4 (2x20μl)",
            "Comp D5 (90μl)",
            "Comp D6 (38μl)",
        ],
    },
    {
        "key": "gabaergic_cells",
        "aliases": ["gabaergic", "gaba", "quick neuron gabaergic", "quick-neuron gabaergic", "ga-mrna-hc-cw50065"],
        "title": "Quick Neuron GABAergic Human iPSC derived Neurons",
        "reagents": [
            "Cryopreserved Cells (GA-mRNA-HC-CW50065)",
            "Comp N (2x840μl)",
            "Comp P (50μl)",
            "Comp G1 (20μl)",
            "Comp G2 (60μl)",
        ],
    },
    {
        "key": "motor_cells",
        "aliases": ["motor", "quick neuron motor", "quick-neuron motor", "mt-sev-hc-cw50065"],
        "title": "Quick Neuron Motor Human iPSC derived Neurons",
        "reagents": [
            "Cryopreserved Cells (MT-SeV-HC-CW50065)",
            "Comp N1 (2x830μl)",
            "Comp A (80μl)",
            "Comp P (50μl)",
            "Comp K (20μl)",
        ],
    },
    {
        "key": "sensory_cells",
        "aliases": ["sensory", "quick neuron sensory", "quick-neuron sensory", "ss-mrna-hc-cw50065"],
        "title": "Quick Neuron Sensory Human iPSC derived Neurons",
        "reagents": [
            "Cryopreserved Cells (SS-mRNA-HC-CW50065)",
            "Comp N (840μl)",
            "Comp P (50μl)",
            "Comp S1 (22μl)",
        ],
    },
    {
        "key": "npc_cells",
        "aliases": ["npc", "neural progenitor", "np-mrna-hc-cw50065"],
        "title": "Quick NPC Neural Progenitor Cells",
        "reagents": [
            "Cryopreserved Cells (NP-mRNA-HC-CW50065)",
        ],
    },
    {
        "key": "astrocyte_cells",
        "aliases": ["astrocyte", "astrocytes", "as-sev-hc-cw50065"],
        "title": "Quick Astrocyte Human iPSC derived Astrocytes",
        "reagents": [
            "Cryopreserved Cells (AS-SeV-HC-CW50065)",
            "Comp P (50μl)",
        ],
    },
    {
        "key": "microglia_cells",
        "aliases": ["microglia", "quick glia", "quick-glia", "mg-sev-hc-cw50065"],
        "title": "Quick Glia Microglia Human iPSC derived Microglia",
        "reagents": [
            "Cryopreserved Cells (MG-SeV-HC-CW50065)",
            "Comp MG1 (55μl)",
            "Comp MG2 (55μl)",
            "Comp MG3 (55μl)",
        ],
    },
    {
        "key": "qn_excitatory",
        "aliases": ["qn-sev-p", "maintenance excitatory", "undiluted 100μl", "comp g1", "comp g2"],
        "title": "Quick Neuron Excitatory Reagent Set",
        "reagents": [
            "QN-SeV-P Undiluted (100μl)",
            "Comp N (3x840μl)",
            "Comp G1 (3x20μl)",
            "Comp G2 (60μl)",
            "Comp P (50μl)",
        ],
    },
    {
        "key": "qn_cholinergic",
        "aliases": ["maintenance cholinergic", "comp a", "qn cholinergic"],
        "title": "Quick Neuron Cholinergic Reagent Set",
        "reagents": [
            "QN-SeV-P Undiluted (100μl)",
            "Comp N1 (3x830μl)",
            "Comp A (80μl)",
            "Comp P (50μl)",
        ],
    },
    {
        "key": "qn_dopaminergic",
        "aliases": ["qnd-mrna-p", "maintenance dopaminergic", "comp d4", "comp d5", "comp d6"],
        "title": "Quick Neuron Dopaminergic Reagent Set",
        "reagents": [
            "QND-mRNA-P (4x33μl)",
            "Comp N (3x840μl)",
            "Comp P (2x50μl)",
            "Comp D4 (4x20μl)",
            "Comp D5 (90μl)",
            "Comp D6 (38μl)",
        ],
    },
    {
        "key": "qn_gabaergic",
        "aliases": ["qng-mrna-p", "maintenance gabaergic"],
        "title": "Quick Neuron GABAergic Reagent Set",
        "reagents": [
            "QNG-mRNA-P (4x33μl)",
            "Comp N (4x840μl)",
            "Comp P (2x50μl)",
            "Comp G1 (3x20μl)",
            "Comp G2 (60μl)",
        ],
    },
    {
        "key": "qn_motor",
        "aliases": ["maintenance motor", "comp k"],
        "title": "Quick Neuron Motor Reagent Set",
        "reagents": [
            "QN-SeV-P Undiluted (100μl)",
            "Comp N1 (3x830μl)",
            "Comp A (80μl)",
            "Comp P (50μl)",
            "Comp K (20μl)",
        ],
    },
    {
        "key": "qn_sensory",
        "aliases": ["qns-mrna-p", "maintenance sensory", "comp s1"],
        "title": "Quick Neuron Sensory Reagent Set",
        "reagents": [
            "QNS-mRNA-P (4x33μl)",
            "Comp N (4x840μl)",
            "Comp S1 (5x22μl)",
            "Comp P (2x50μl)",
        ],
    },
    {
        "key": "qga",
        "aliases": ["qga-sev", "ga1", "ga2"],
        "title": "Quick Astrocyte Reagent Set",
        "reagents": [
            "QGA-SeV Undiluted (100μl)",
            "Comp N1 (2x830μl)",
            "Comp GA1 (65μl)",
            "Comp GA2 (65μl)",
        ],
    },
    {
        "key": "qms",
        "aliases": ["qms-sev", "mesendoderm", "mesendoderm rna-p", "component p"],
        "title": "Quick Mesendoderm Reagent Set",
        "reagents": [
            "QMS-SeV (110 μl)",
            "Mesendoderm RNA-P (36 μl)",
            "Component P (50 μl)",
        ],
    },
]


def parse_json_response(raw_text):
    raw_text = raw_text.strip()

    if raw_text.startswith("```"):
        raw_text = raw_text.replace("```json", "").replace("```", "").strip()

    start = raw_text.find("{")
    end = raw_text.rfind("}")

    if start != -1 and end != -1 and end > start:
        raw_text = raw_text[start:end + 1]

    return json.loads(raw_text)


def extract_text_from_pdf(uploaded_file):
    try:
        uploaded_file.seek(0)
        reader = PdfReader(uploaded_file)
        parts = []

        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                parts.append(page_text)

        return "\n".join(parts).strip()
    except Exception:
        return ""


def extract_text_from_pdf_bytes(file_bytes):
    try:
        buffer = io.BytesIO(file_bytes)
        reader = PdfReader(buffer)
        parts = []

        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                parts.append(page_text)

        return "\n".join(parts).strip()
    except Exception:
        return ""


def extract_text_from_txt(uploaded_file):
    uploaded_file.seek(0)
    return StringIO(uploaded_file.getvalue().decode("utf-8")).read().strip()


def extract_file_text(uploaded_file):
    if uploaded_file.type == "application/pdf":
        return extract_text_from_pdf(uploaded_file)

    return extract_text_from_txt(uploaded_file)


def normalize_for_match(value):
    if value is None:
        return ""

    text = str(value).lower().strip()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = " ".join(text.split())

    return text


def is_shipping_line(description):
    text = normalize_for_match(description)
    shipping_words = ["shipping", "delivery", "handling", "freight"]

    return any(word in text for word in shipping_words)


def standardize_component_name(text):
    if text is None:
        return ""

    text = str(text).strip()
    text = re.sub(r"\bComp\b", "Component", text)

    return text


def build_ln2_storage_line(product_names):
    names = [name for name in product_names if name]

    if not names:
        product_text = "the purchased product"
    elif len(names) == 1:
        product_text = names[0]
    elif len(names) == 2:
        product_text = f"{names[0]} and {names[1]}"
    else:
        product_text = ", ".join(names[:-1]) + f", and {names[-1]}"

    return f"Please store {product_text} in LN2 storage upon receipt."


@st.cache_data(show_spinner=False)
def extract_quote_json_cached(file_bytes, mime_type):
    if mime_type == "application/pdf":
        pdf_part = types.Part.from_bytes(data=file_bytes, mime_type="application/pdf")

        prompt = """
Extract only the fields needed for invoicing and delivery note creation from this quote.

Important rules:
1. Read all visible text from all pages.
2. Extract quote_number, client_company_name, issue_date, currency, shipping_fee, total_amount.
3. Extract only product lines needed for shipping.
4. Exclude shipping, delivery, handling, or freight lines from items.
5. Capture each quoted product description, sku if visible, and quantity exactly as shown.
6. Return only JSON.

Use this schema:
{
  "quote_number": "",
  "issue_date": "",
  "client_company_name": "",
  "currency": "",
  "shipping_fee": "",
  "total_amount": "",
  "items": [
    {
      "description": "",
      "sku": "",
      "quantity": ""
    }
  ]
}
"""

        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=[pdf_part, prompt],
        )

        raw_text = response.text.strip()

    else:
        text = file_bytes.decode("utf-8")

        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=f"""
Extract only the fields needed for invoicing and delivery note creation from this quote.

Important rules:
1. Extract quote_number, client_company_name, issue_date, currency, shipping_fee, total_amount.
2. Extract only product lines needed for shipping.
3. Exclude shipping, delivery, handling, or freight lines from items.
4. Capture each quoted product description, sku if visible, and quantity exactly as shown.
5. Return only JSON.

Use this schema:
{{
  "quote_number": "",
  "issue_date": "",
  "client_company_name": "",
  "currency": "",
  "shipping_fee": "",
  "total_amount": "",
  "items": [
    {{
      "description": "",
      "sku": "",
      "quantity": ""
    }}
  ]
}}

Quote Text:
{text}
""",
        )

        raw_text = response.text.strip()

    return parse_json_response(raw_text)


@st.cache_data(show_spinner=False)
def extract_po_json_cached(file_bytes, mime_type):
    if mime_type == "application/pdf":
        po_text = extract_text_from_pdf_bytes(file_bytes)
    else:
        po_text = file_bytes.decode("utf-8")

    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=f"""
Extract only the fields needed for invoicing and delivery note creation from this purchase order.

Important rules:
1. Extract po_number, client_company_name, supplier_name, payment_terms, ship_to_address, bill_to_address, ap_accounting_email, currency, shipping_fee, total_amount.
2. Return only JSON.

Use this schema:
{{
  "po_number": "",
  "client_company_name": "",
  "supplier_name": "",
  "payment_terms": "",
  "ship_to_address": "",
  "bill_to_address": "",
  "ap_accounting_email": "",
  "currency": "",
  "shipping_fee": "",
  "total_amount": ""
}}

Purchase Order Text:
{po_text}
""",
    )

    raw_text = response.text.strip()

    return parse_json_response(raw_text)


def find_catalog_match(description, sku=""):
    haystack = f"{description} {sku}".lower()
    best_match = None
    best_score = 0

    for entry in PRODUCT_CATALOG:
        score = 0

        for alias in entry["aliases"]:
            alias_lower = alias.lower()

            if alias_lower in haystack:
                score = max(score, len(alias_lower))

        if score > best_score:
            best_score = score
            best_match = entry

    return best_match


def build_invoice_details(quote_data, po_data):
    ship_to = po_data.get("ship_to_address", "")
    bill_to = po_data.get("bill_to_address", "")
    billing_email = po_data.get("ap_accounting_email", "")
    po_number = po_data.get("po_number", "")
    client_name = po_data.get("client_company_name", "") or quote_data.get("client_company_name", "")
    currency = po_data.get("currency", "") or quote_data.get("currency", "")
    total_amount = po_data.get("total_amount", "") or quote_data.get("total_amount", "")
    shipping_fee = po_data.get("shipping_fee", "") or quote_data.get("shipping_fee", "")

    quote_items = quote_data.get("items", [])
    quote_items = [item for item in quote_items if not is_shipping_line(item.get("description", ""))]

    products = []

    for item in quote_items:
        products.append({
            "description": item.get("description", ""),
            "sku": item.get("sku", ""),
            "quantity": item.get("quantity", "")
        })

    return {
        "po_number": po_number,
        "client_company_name": client_name,
        "shipping_address": ship_to,
        "billing_address": bill_to,
        "billing_email": billing_email,
        "currency": currency,
        "total_amount": total_amount,
        "shipping_fee": shipping_fee,
        "products": products
    }


def build_po_checks(po_data):
    payment_terms = str(po_data.get("payment_terms", "")).strip()
    shipping_fee = str(po_data.get("shipping_fee", "")).strip()
    supplier_name = str(po_data.get("supplier_name", "")).strip()

    payment_terms_ok = "net 30" in payment_terms.lower()
    shipping_fee_specified = shipping_fee != ""
    supplier_ok = "ricoh biosciences" in supplier_name.lower()

    rows = [
        {
            "Check": "Payment terms specify Net 30",
            "Yes or No": "Yes" if payment_terms_ok else "No",
            "Details": payment_terms if payment_terms else "Missing"
        },
        {
            "Check": "Shipping fee specified in PO",
            "Yes or No": "Yes" if shipping_fee_specified else "No",
            "Details": shipping_fee if shipping_fee else "Missing"
        },
        {
            "Check": "Supplier correctly states Ricoh Biosciences",
            "Yes or No": "Yes" if supplier_ok else "No",
            "Details": supplier_name if supplier_name else "Missing"
        },
    ]

    return pd.DataFrame(rows)


def build_delivery_note_items_from_quote(quote_data):
    quote_items = quote_data.get("items", [])
    quote_items = [item for item in quote_items if not is_shipping_line(item.get("description", ""))]

    delivery_items = []

    for item in quote_items:
        desc = item.get("description", "")
        sku = item.get("sku", "")
        match = find_catalog_match(desc, sku)

        if match is not None:
            official_title = match["title"]
            contents = [standardize_component_name(r) for r in match["reagents"]]
        else:
            official_title = desc
            contents = []

        delivery_items.append({
            "description": desc,
            "sku": sku,
            "quantity": item.get("quantity", ""),
            "catalog_match": match,
            "display_title": official_title,
            "contents": contents
        })

    return delivery_items


def build_delivery_note_preview_df(delivery_items):
    rows = []

    for idx, item in enumerate(delivery_items, start=1):
        rows.append({
            "item_index": idx,
            "quoted_product": item["description"],
            "quantity": item["quantity"],
            "delivery_title": item.get("display_title", ""),
            "catalog_matched": "" if item["catalog_match"] is None else item["catalog_match"]["title"],
            "contents_count": len(item.get("contents", []))
        })

    return pd.DataFrame(rows)


def choose_template_sheet_name(wb):
    for candidate in ["260306", "251208", "251015", "260330"]:
        if candidate in wb.sheetnames:
            return candidate

    return wb.sheetnames[0]


def get_writable_cell(ws, row, col):
    cell = ws.cell(row, col)

    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(merged_range.min_row, merged_range.min_col)

    return cell


def set_cell_value_safe(ws, cell_ref=None, row=None, col=None, value=""):
    if cell_ref is not None:
        row = ws[cell_ref].row
        col = ws[cell_ref].column

    target = get_writable_cell(ws, row, col)
    target.value = value


def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)

        if source.has_style:
            target._style = copy(source._style)

        if source.number_format:
            target.number_format = source.number_format

        if source.alignment:
            target.alignment = copy(source.alignment)

        if source.font:
            target.font = copy(source.font)

        if source.fill:
            target.fill = copy(source.fill)

        if source.border:
            target.border = copy(source.border)

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def populate_delivery_note_template(workbook_path, output_path, po_number, delivery_items):
    wb = load_workbook(workbook_path)
    ws = wb.active

    # -------------------
    # PO NUMBER
    # -------------------
    ws["G16"] = f"PO# {po_number}" if po_number else "PO#"

    matched_items = [item for item in delivery_items if item["catalog_match"] is not None]

    # -------------------
    # PRODUCT 1
    # -------------------
    if len(matched_items) >= 1:
        item1 = matched_items[0]

        ws["G18"] = item1["display_title"]  # Product name
        ws["G20"] = item1["display_title"]  # Full name

        start_row = 21
        for i, comp in enumerate(item1["contents"]):
            ws[f"G{start_row + i}"] = standardize_component_name(comp)

    # -------------------
    # PRODUCT 2 (FIXED ROW)
    # -------------------
    if len(matched_items) >= 2:
        item2 = matched_items[1]

        ws["G29"] = item2["display_title"]  # <-- FIXED HERE
        ws["G31"] = item2["display_title"]

        start_row = 32
        for i, comp in enumerate(item2["contents"]):
            ws[f"G{start_row + i}"] = standardize_component_name(comp)

    # -------------------
    # STORAGE LINE
    # -------------------
    product_names = [item["display_title"] for item in matched_items]

    if len(product_names) == 1:
        storage_text = f'Please store {product_names[0]} in LN2 storage upon receipt.'
    elif len(product_names) == 2:
        storage_text = f'Please store {product_names[0]} and {product_names[1]} in LN2 storage upon receipt.'
    else:
        storage_text = f'Please store all products in LN2 storage upon receipt.'

    ws["A45"] = storage_text

    wb.save(output_path)


def generate_delivery_note_file(po_data, delivery_items, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    matched_items = [item for item in delivery_items if item["catalog_match"] is not None]

    if not matched_items:
        return None

    po_number = po_data.get("po_number", "PO")
    safe_po = re.sub(r"[^A-Za-z0-9 _()]+", "", po_number).strip() or "PO"

    filename = f"{safe_po}_delivery_note.xlsx"
    output_path = os.path.join(output_dir, filename)

    populate_delivery_note_template(
        TEMPLATE_PATH,
        output_path,
        po_data.get("po_number", ""),
        matched_items
    )

    return output_path


def ask_question_about_docs(question, quote_data, po_data, delivery_items):
    prompt = f"""
You are answering questions about a Quote, a Purchase Order, and the current delivery note draft.

Use only the information provided below.
If the answer is not available, say that clearly.
Be concise and specific.

Quote data:
{json.dumps(quote_data, indent=2)}

PO data:
{json.dumps(po_data, indent=2)}

Current delivery note draft:
{json.dumps(delivery_items, indent=2)}

User question:
{question}
"""

    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=prompt,
    )

    return response.text.strip()


def interpret_delivery_note_edit(request_text, delivery_items, po_data):
    prompt = f"""
You convert user edit requests into structured JSON edits for a delivery note draft.

Only use these allowed action types:
update_po_number
update_item_title
add_content
remove_content
replace_content

Rules:
1. item_index is 1 based.
2. Return only JSON.
3. If the request is unclear, return an empty actions list.
4. Do not invent items that are not in the draft.
5. If product title changes, storage line will be regenerated automatically.

Current delivery note draft:
{json.dumps(delivery_items, indent=2)}

Current PO data:
{json.dumps(po_data, indent=2)}

Return this schema:
{{
  "actions": [
    {{
      "action_type": "",
      "item_index": 0,
      "value": "",
      "old_value": "",
      "new_value": ""
    }}
  ]
}}

User request:
{request_text}
"""

    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=prompt,
    )

    return parse_json_response(response.text)


def apply_delivery_note_edits(delivery_items, po_data, edit_plan):
    updated_items = deepcopy(delivery_items)
    updated_po_data = deepcopy(po_data)
    actions = edit_plan.get("actions", [])

    applied = []

    for action in actions:
        action_type = action.get("action_type", "")
        item_index = action.get("item_index", 0)

        if action_type == "update_po_number":
            updated_po_data["po_number"] = action.get("value", "")
            applied.append(f'Updated PO number to "{action.get("value", "")}"')
            continue

        if item_index < 1 or item_index > len(updated_items):
            continue

        item = updated_items[item_index - 1]

        if action_type == "update_item_title":
            item["display_title"] = action.get("value", "")
            applied.append(f"Updated item {item_index} title")

        elif action_type == "add_content":
            value = standardize_component_name(action.get("value", ""))

            if value and value not in item["contents"]:
                item["contents"].append(value)
                applied.append(f"Added content line to item {item_index}")

        elif action_type == "remove_content":
            value = standardize_component_name(action.get("value", ""))

            if value in item["contents"]:
                item["contents"].remove(value)
                applied.append(f"Removed content line from item {item_index}")

        elif action_type == "replace_content":
            old_value = standardize_component_name(action.get("old_value", ""))
            new_value = standardize_component_name(action.get("new_value", ""))

            if old_value in item["contents"]:
                item["contents"] = [new_value if r == old_value else r for r in item["contents"]]
                applied.append(f"Replaced content line on item {item_index}")

    return updated_items, updated_po_data, applied


st.sidebar.header("Upload Files")
quote_file = st.sidebar.file_uploader("Upload Quote", type=["pdf", "txt"], key="quote")
po_file = st.sidebar.file_uploader("Upload PO", type=["pdf", "txt"], key="po")

if quote_file is not None and po_file is not None:
    quote_preview_text = extract_file_text(quote_file)
    po_preview_text = extract_file_text(po_file)

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Quote Text Preview")
        st.text_area("Quote preview", quote_preview_text, height=200)

    with col2:
        st.subheader("PO Text Preview")
        st.text_area("PO preview", po_preview_text, height=200)

    if st.button("Build Invoice and Delivery Note"):
        try:
            quote_bytes = quote_file.getvalue()
            po_bytes = po_file.getvalue()

            with st.spinner("Reading quote..."):
                quote_data = extract_quote_json_cached(quote_bytes, quote_file.type)

            with st.spinner("Reading PO..."):
                po_data = extract_po_json_cached(po_bytes, po_file.type)

            delivery_items = build_delivery_note_items_from_quote(quote_data)

            st.session_state["quote_data"] = quote_data
            st.session_state["po_data"] = po_data
            st.session_state["delivery_items"] = delivery_items
            st.session_state["qa_answer"] = ""
            st.session_state["edit_log"] = []

        except Exception as e:
            st.error(f"Error while extracting files: {e}")

if "quote_data" in st.session_state and "po_data" in st.session_state and "delivery_items" in st.session_state:
    quote_data = st.session_state["quote_data"]
    po_data = st.session_state["po_data"]
    delivery_items = st.session_state["delivery_items"]

    st.subheader("Extracted Quote JSON")
    st.json(quote_data)

    st.subheader("Extracted PO JSON")
    st.json(po_data)

    st.subheader("PO Checks")
    st.dataframe(build_po_checks(po_data), use_container_width=True)

    invoice_details = build_invoice_details(quote_data, po_data)

    st.subheader("Invoice Details")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**PO Number**")
        st.write(invoice_details["po_number"] or "Missing")

        st.markdown("**Client Company Name**")
        st.write(invoice_details["client_company_name"] or "Missing")

        st.markdown("**Shipping Address**")
        st.write(invoice_details["shipping_address"] or "Missing")

        st.markdown("**Billing Email**")
        st.write(invoice_details["billing_email"] or "Missing")

    with col2:
        st.markdown("**Billing Address**")
        st.write(invoice_details["billing_address"] or "Missing")

        st.markdown("**Currency**")
        st.write(invoice_details["currency"] or "Missing")

        st.markdown("**Total Amount**")
        st.write(invoice_details["total_amount"] or "Missing")

        st.markdown("**Shipping Fee**")
        st.write(invoice_details["shipping_fee"] or "Missing")

    st.markdown("**Products for Invoice**")

    if invoice_details["products"]:
        st.dataframe(pd.DataFrame(invoice_details["products"]), use_container_width=True)
    else:
        st.write("No products found on quote.")

    st.subheader("Ask About the PO or Quote")
    qa_question = st.text_input("Ask a question about the Quote, PO, or current delivery note draft")

    if st.button("Ask Gemini"):
        try:
            with st.spinner("Thinking..."):
                answer = ask_question_about_docs(qa_question, quote_data, po_data, delivery_items)

            st.session_state["qa_answer"] = answer

        except Exception as e:
            st.error(f"Question error: {e}")

    if st.session_state.get("qa_answer"):
        st.markdown("**Answer**")
        st.write(st.session_state["qa_answer"])

    st.subheader("Delivery Note Draft")
    st.dataframe(build_delivery_note_preview_df(delivery_items), use_container_width=True)

    edit_request = st.text_input("Request an edit to the delivery note draft")

    if st.button("Apply Delivery Note Edit"):
        try:
            with st.spinner("Applying edit..."):
                edit_plan = interpret_delivery_note_edit(edit_request, delivery_items, po_data)
                updated_items, updated_po_data, applied = apply_delivery_note_edits(delivery_items, po_data, edit_plan)

            st.session_state["delivery_items"] = updated_items
            st.session_state["po_data"] = updated_po_data
            st.session_state["edit_log"] = st.session_state.get("edit_log", []) + applied

            if not applied:
                st.info("No clear edit was applied.")

        except Exception as e:
            st.error(f"Edit error: {e}")

    if st.session_state.get("edit_log"):
        st.markdown("**Applied edits**")

        for line in st.session_state["edit_log"]:
            st.write(f"• {line}")

    unmatched = [item for item in delivery_items if item["catalog_match"] is None]

    for item in unmatched:
        st.warning(f'No catalog match found for quoted item: {item["description"]}')

    if not os.path.exists(TEMPLATE_PATH):
        st.warning(f'Put "{TEMPLATE_PATH}" in the same folder as app.py to generate the delivery note.')

    matched_count = sum(1 for item in delivery_items if item["catalog_match"] is not None)

    if matched_count > 0 and os.path.exists(TEMPLATE_PATH):
        if st.button("Generate Delivery Note XLSX"):
            try:
                with tempfile.TemporaryDirectory() as temp_dir:
                    delivery_note_path = generate_delivery_note_file(
                        st.session_state["po_data"],
                        st.session_state["delivery_items"],
                        temp_dir
                    )

                    if not delivery_note_path:
                        st.error("No delivery note was generated because no quoted products matched the catalog.")
                        st.stop()

                    with open(delivery_note_path, "rb") as f:
                        st.success("Delivery note generated.")
                        st.download_button(
                            label="Download Delivery Note XLSX",
                            data=f.read(),
                            file_name=os.path.basename(delivery_note_path),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            except Exception as e:
                st.error(f"Error while generating delivery note: {e}")

    st.subheader("Reminders")
    st.write("1. Tag admin")
    st.write("2. Update HubSpot")

else:
    st.info("Please upload both a Quote and a PO.")